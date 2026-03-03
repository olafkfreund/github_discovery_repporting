from __future__ import annotations

"""Unit tests for Excel, Markdown, and Zip report export renderers.

These tests exercise the three synchronous renderer classes directly without
requiring a running database, HTTP server, or async event loop.  All file
I/O uses pytest's ``tmp_path`` fixture so nothing is written to the real
filesystem.
"""

import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.reports.excel import ExcelRenderer
from backend.reports.markdown import MarkdownRenderer, _md_table
from backend.reports.zip_bundler import ZipBundler

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def sample_report_data() -> dict:
    """Return a minimal but complete report data dict.

    The structure mirrors what ``ReportGenerator._build_report_data()``
    produces so that each renderer receives realistic input.
    """
    finding_cicd = {
        "check_id": "CICD-001",
        "check_name": "CI Pipeline Exists",
        "category": "CI/CD Pipeline",
        "severity": "high",
        "status": "passed",
        "detail": "At least one CI workflow was detected.",
        "score": 100,
    }
    finding_iam = {
        "check_id": "IAM-003",
        "check_name": "MFA Enforced",
        "category": "Identity & Access",
        "severity": "critical",
        "status": "failed",
        "detail": "MFA is not enforced for all members.",
        "score": 0,
    }
    finding_pipe = {
        "check_id": "IAM-004",
        "check_name": "SSO Integration",
        "category": "Identity & Access",
        "severity": "medium",
        "status": "warning",
        "detail": "SSO could not be verified via API.",
        "score": 50,
    }

    return {
        # --- Metadata ---
        "report_title": "DevOps Assessment Report",
        "customer_name": "Acme Corporation",
        "org_name": "acme-org",
        "scan_id": "d4e5f6a7-b8c9-4000-8000-000000000001",
        "generated_at": "2026-03-03T12:00:00Z",
        "platform": "github",
        "platform_display_name": "GitHub",
        # --- Scores ---
        "overall_score": 72.5,
        "dora_level": "Medium",
        # --- Category data ---
        "category_scores": {
            "CI/CD Pipeline": 88.0,
            "Identity & Access": 45.0,
        },
        "category_narratives": [
            {
                "category": "CI/CD Pipeline",
                "strengths": ["All repos have CI workflows.", "High pipeline success rate."],
                "weaknesses": ["No deployment gating on staging."],
            },
            {
                "category": "Identity & Access",
                "strengths": ["Admin list is small."],
                "weaknesses": ["MFA not enforced.", "No SAML SSO configured."],
            },
        ],
        # --- Narrative ---
        "executive_summary": (
            "The organisation has a solid CI/CD foundation but faces significant "
            "identity and access management gaps."
        ),
        "overall_maturity": "Developing",
        "risk_highlights": [
            "MFA not enforced for all members.",
            "No secrets scanning baseline established.",
        ],
        # --- Recommendations ---
        "recommendations": [
            {
                "priority": "Critical",
                "title": "Enforce MFA Across All Members",
                "category": "Identity & Access",
                "effort": "Low",
                "impact": "High",
                "description": "Enable mandatory MFA at the organisation level.",
                "related_check_ids": ["IAM-003"],
            },
            {
                "priority": "High",
                "title": "Add Secrets Scanning Baseline",
                "category": "Secrets Management",
                "effort": "Medium",
                "impact": "High",
                "description": "Enable GitHub Advanced Security secret scanning.",
                "related_check_ids": ["SEC-001", "SEC-002"],
            },
        ],
        # --- Benchmarks ---
        "benchmark_comparisons": [
            {
                "framework": "OpenSSF Scorecard",
                "level": "3 / 10",
                "summary": "Below average; missing branch protection and signing.",
            },
            {
                "framework": "DORA",
                "level": "Medium",
                "summary": "Deployment frequency and lead time are acceptable.",
            },
        ],
        # --- Findings ---
        "all_findings": [finding_cicd, finding_iam, finding_pipe],
        "findings_by_category": {
            "CI/CD Pipeline": [finding_cicd],
            "Identity & Access": [finding_iam, finding_pipe],
        },
    }


# ---------------------------------------------------------------------------
# ExcelRenderer tests
# ---------------------------------------------------------------------------


class TestExcelRenderer:
    """Unit tests for :class:`~backend.reports.excel.ExcelRenderer`."""

    def test_generate_excel_creates_file(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """generate_excel() must write a file at the requested output_path."""
        output = tmp_path / "report.xlsx"
        renderer = ExcelRenderer()

        result = renderer.generate_excel(sample_report_data, output)

        assert result == output
        assert output.exists()
        assert output.suffix == ".xlsx"

    def test_generate_excel_has_six_sheets(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """The generated workbook must contain exactly six sheets."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        assert len(wb.sheetnames) == 6

    def test_generate_excel_sheet_names(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """The six sheets must use the expected tab names in order."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        assert wb.sheetnames == [
            "Cover",
            "Executive Summary",
            "Category Scores",
            "Recommendations",
            "Benchmarks",
            "All Findings",
        ]

    def test_generate_excel_cover_sheet_customer_name(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """Cover sheet must contain the customer name in the Value column."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        ws = wb["Cover"]
        values = [str(ws.cell(row=r, column=2).value) for r in range(2, ws.max_row + 1)]
        assert "Acme Corporation" in values

    def test_generate_excel_cover_sheet_overall_score(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """Cover sheet Value column must include the formatted overall score."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        ws = wb["Cover"]
        values = [str(ws.cell(row=r, column=2).value) for r in range(2, ws.max_row + 1)]
        assert "72.5%" in values

    def test_generate_excel_cover_sheet_platform_display_name(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """Cover sheet must display the platform_display_name, not the raw slug."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        ws = wb["Cover"]
        values = [str(ws.cell(row=r, column=2).value) for r in range(2, ws.max_row + 1)]
        assert "GitHub" in values

    def test_generate_excel_findings_sheet_headers(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """'All Findings' sheet header row must contain all seven expected column names."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        ws = wb["All Findings"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == [
            "Check ID",
            "Name",
            "Category",
            "Severity",
            "Status",
            "Detail",
            "Score",
        ]

    def test_generate_excel_findings_sheet_data_rows(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """'All Findings' sheet must contain one data row per finding."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        ws = wb["All Findings"]
        # Row 1 is the header; remaining rows are data.
        data_rows = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None
        ]
        expected_ids = {"CICD-001", "IAM-003", "IAM-004"}
        assert set(data_rows) == expected_ids

    def test_generate_excel_recommendations_sheet_data(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """'Recommendations' sheet must have two data rows matching fixture data."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        ws = wb["Recommendations"]
        titles = [
            ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=2).value is not None
        ]
        assert "Enforce MFA Across All Members" in titles
        assert "Add Secrets Scanning Baseline" in titles

    def test_generate_excel_creates_parent_dirs(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """generate_excel() must create intermediate directories if they are absent."""
        output = tmp_path / "nested" / "deep" / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)
        assert output.exists()

    def test_generate_excel_category_scores_sheet_has_data(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """'Category Scores' sheet must include category names from the fixture."""
        output = tmp_path / "report.xlsx"
        ExcelRenderer().generate_excel(sample_report_data, output)

        wb = load_workbook(output)
        ws = wb["Category Scores"]
        categories = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None
        ]
        assert "CI/CD Pipeline" in categories
        assert "Identity & Access" in categories


# ---------------------------------------------------------------------------
# MarkdownRenderer tests
# ---------------------------------------------------------------------------


class TestMarkdownRenderer:
    """Unit tests for :class:`~backend.reports.markdown.MarkdownRenderer`."""

    def test_generate_markdown_creates_six_files(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """generate_markdown() must produce exactly six .md files in output_dir."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        md_files = sorted(out_dir.glob("*.md"))
        assert len(md_files) == 6

    def test_generate_markdown_expected_filenames(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """The six files must use the exact filename conventions."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        names = {f.name for f in out_dir.glob("*.md")}
        assert names == {
            "00-cover.md",
            "01-executive-summary.md",
            "02-category-detail.md",
            "03-recommendations.md",
            "04-benchmarks.md",
            "05-appendix-findings.md",
        }

    def test_generate_markdown_returns_output_dir(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """generate_markdown() must return the output directory path."""
        out_dir = tmp_path / "md_out"
        result = MarkdownRenderer().generate_markdown(sample_report_data, out_dir)
        assert result == out_dir

    def test_markdown_cover_contains_customer_name(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """00-cover.md must include the customer name."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        content = (out_dir / "00-cover.md").read_text(encoding="utf-8")
        assert "Acme Corporation" in content

    def test_markdown_cover_contains_overall_score(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """00-cover.md must display the formatted overall score."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        content = (out_dir / "00-cover.md").read_text(encoding="utf-8")
        assert "72.5%" in content

    def test_markdown_cover_contains_dora_level(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """00-cover.md must include the DORA level value."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        content = (out_dir / "00-cover.md").read_text(encoding="utf-8")
        assert "Medium" in content

    def test_markdown_recommendations_table_uses_pipe_syntax(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """03-recommendations.md must contain a Markdown pipe table."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        content = (out_dir / "03-recommendations.md").read_text(encoding="utf-8")
        # Every pipe table row starts and ends with '|'
        table_lines = [ln for ln in content.splitlines() if ln.startswith("|")]
        assert len(table_lines) >= 3  # header + separator + at least one data row

    def test_markdown_recommendations_table_contains_recommendation_titles(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """03-recommendations.md summary table must list recommendation titles."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        content = (out_dir / "03-recommendations.md").read_text(encoding="utf-8")
        assert "Enforce MFA Across All Members" in content
        assert "Add Secrets Scanning Baseline" in content

    def test_markdown_escapes_pipes_in_cells(
        self, tmp_path: Path
    ) -> None:
        """Pipe characters inside cell values must be escaped as '\\|'."""
        data: dict = {
            "report_title": "Test",
            "customer_name": "Foo|Bar",  # pipe in value
            "org_name": "org",
            "scan_id": "1",
            "generated_at": "now",
            "platform": "github",
            "platform_display_name": "GitHub",
            "overall_score": 0.0,
            "dora_level": "Low",
            "category_scores": {"A|B": 50.0},
            "category_narratives": [
                {"category": "A|B", "strengths": [], "weaknesses": []}
            ],
            "executive_summary": "summary",
            "overall_maturity": "low",
            "risk_highlights": [],
            "recommendations": [
                {
                    "priority": "High",
                    "title": "Fix|Issue",
                    "category": "Cat|Sub",
                    "effort": "Low",
                    "impact": "High",
                    "description": "desc",
                    "related_check_ids": [],
                }
            ],
            "benchmark_comparisons": [],
            "all_findings": [
                {
                    "check_id": "X-001",
                    "check_name": "Check|Name",
                    "category": "Cat",
                    "severity": "low",
                    "status": "passed",
                    "detail": "detail|info",
                    "score": 100,
                }
            ],
            "findings_by_category": {
                "A|B": [
                    {
                        "check_id": "X-001",
                        "check_name": "Check|Name",
                        "category": "A|B",
                        "severity": "low",
                        "status": "passed",
                        "detail": "detail|info",
                        "score": 100,
                    }
                ]
            },
        }
        out_dir = tmp_path / "md_escape"
        MarkdownRenderer().generate_markdown(data, out_dir)

        # Recommendations table contains a cell with a pipe — must be escaped.
        recs_content = (out_dir / "03-recommendations.md").read_text(encoding="utf-8")
        assert "Fix\\|Issue" in recs_content

        # Findings appendix contains a cell with a pipe — must be escaped.
        findings_content = (out_dir / "05-appendix-findings.md").read_text(encoding="utf-8")
        assert "Check\\|Name" in findings_content

    def test_markdown_executive_summary_contains_risk_highlights(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """01-executive-summary.md must list all risk highlights as bullet points."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        content = (out_dir / "01-executive-summary.md").read_text(encoding="utf-8")
        assert "- MFA not enforced for all members." in content
        assert "- No secrets scanning baseline established." in content

    def test_markdown_creates_output_dir_if_absent(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """generate_markdown() must create the output directory when it does not exist."""
        out_dir = tmp_path / "new" / "nested" / "dir"
        assert not out_dir.exists()
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)
        assert out_dir.is_dir()

    def test_markdown_appendix_findings_contains_check_ids(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """05-appendix-findings.md must include the check IDs from all_findings."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        content = (out_dir / "05-appendix-findings.md").read_text(encoding="utf-8")
        assert "CICD-001" in content
        assert "IAM-003" in content
        assert "IAM-004" in content

    def test_markdown_benchmarks_table_present(
        self, tmp_path: Path, sample_report_data: dict
    ) -> None:
        """04-benchmarks.md must contain framework names from benchmark_comparisons."""
        out_dir = tmp_path / "md_out"
        MarkdownRenderer().generate_markdown(sample_report_data, out_dir)

        content = (out_dir / "04-benchmarks.md").read_text(encoding="utf-8")
        assert "OpenSSF Scorecard" in content
        assert "DORA" in content


# ---------------------------------------------------------------------------
# ZipBundler tests
# ---------------------------------------------------------------------------


class TestZipBundler:
    """Unit tests for :class:`~backend.reports.zip_bundler.ZipBundler`."""

    def test_create_zip_from_files_creates_archive(self, tmp_path: Path) -> None:
        """create_zip() must write a valid .zip file at the requested output_path."""
        file_a = tmp_path / "a.txt"
        file_b = tmp_path / "b.txt"
        file_a.write_text("alpha")
        file_b.write_text("beta")

        output = tmp_path / "bundle.zip"
        result = ZipBundler().create_zip(
            [(file_a, "a.txt"), (file_b, "b.txt")],
            output,
        )

        assert result == output
        assert output.exists()
        assert zipfile.is_zipfile(output)

    def test_create_zip_contains_expected_archive_names(self, tmp_path: Path) -> None:
        """The zip archive must contain entries with the specified archive names."""
        file_a = tmp_path / "alpha.txt"
        file_b = tmp_path / "beta.md"
        file_a.write_text("content-a")
        file_b.write_text("content-b")

        output = tmp_path / "bundle.zip"
        ZipBundler().create_zip(
            [(file_a, "reports/alpha.txt"), (file_b, "reports/beta.md")],
            output,
        )

        with zipfile.ZipFile(output) as zf:
            names = zf.namelist()

        assert "reports/alpha.txt" in names
        assert "reports/beta.md" in names

    def test_create_zip_file_content_preserved(self, tmp_path: Path) -> None:
        """File contents must be preserved verbatim inside the archive."""
        source = tmp_path / "data.txt"
        source.write_text("hello-from-test")

        output = tmp_path / "bundle.zip"
        ZipBundler().create_zip([(source, "data.txt")], output)

        with zipfile.ZipFile(output) as zf:
            recovered = zf.read("data.txt").decode("utf-8")

        assert recovered == "hello-from-test"

    def test_create_zip_from_directory_archives_all_children(self, tmp_path: Path) -> None:
        """When a directory is passed, all files inside it must appear in the archive."""
        src_dir = tmp_path / "markdown"
        src_dir.mkdir()
        (src_dir / "00-cover.md").write_text("# Cover")
        (src_dir / "01-exec.md").write_text("# Exec")
        (src_dir / "02-cat.md").write_text("# Cat")

        output = tmp_path / "bundle.zip"
        ZipBundler().create_zip([(src_dir, "markdown")], output)

        with zipfile.ZipFile(output) as zf:
            names = set(zf.namelist())

        assert "markdown/00-cover.md" in names
        assert "markdown/01-exec.md" in names
        assert "markdown/02-cat.md" in names

    def test_create_zip_creates_parent_directories(self, tmp_path: Path) -> None:
        """create_zip() must create intermediate directories for the output path."""
        source = tmp_path / "file.txt"
        source.write_text("x")

        output = tmp_path / "nested" / "deep" / "bundle.zip"
        ZipBundler().create_zip([(source, "file.txt")], output)

        assert output.exists()

    def test_create_zip_skips_nonexistent_source(self, tmp_path: Path) -> None:
        """A source path that does not exist must be silently skipped."""
        ghost = tmp_path / "ghost.txt"  # never created
        output = tmp_path / "bundle.zip"

        ZipBundler().create_zip([(ghost, "ghost.txt")], output)

        with zipfile.ZipFile(output) as zf:
            assert "ghost.txt" not in zf.namelist()

    def test_create_zip_returns_output_path(self, tmp_path: Path) -> None:
        """create_zip() must return the output_path it was given."""
        source = tmp_path / "x.txt"
        source.write_text("x")
        output = tmp_path / "out.zip"

        result = ZipBundler().create_zip([(source, "x.txt")], output)

        assert result == output

    def test_create_zip_multiple_files_count(self, tmp_path: Path) -> None:
        """The archive must contain exactly as many entries as valid source files."""
        files = []
        for i in range(5):
            p = tmp_path / f"f{i}.txt"
            p.write_text(f"content-{i}")
            files.append((p, f"f{i}.txt"))

        output = tmp_path / "multi.zip"
        ZipBundler().create_zip(files, output)

        with zipfile.ZipFile(output) as zf:
            assert len(zf.namelist()) == 5


# ---------------------------------------------------------------------------
# _md_table helper tests
# ---------------------------------------------------------------------------


class TestMdTable:
    """Unit tests for the :func:`~backend.reports.markdown._md_table` helper."""

    def test_md_table_produces_pipe_rows(self) -> None:
        """Every line in the output must begin and end with a pipe character."""
        result = _md_table(["A", "B"], [["1", "2"], ["3", "4"]])
        for line in result.splitlines():
            assert line.startswith("|"), f"Line does not start with |: {line!r}"
            assert line.endswith("|"), f"Line does not end with |: {line!r}"

    def test_md_table_header_and_separator_present(self) -> None:
        """Output must contain the header row and the separator row."""
        result = _md_table(["Col1", "Col2"], [["val1", "val2"]])
        lines = result.splitlines()
        assert lines[0] == "| Col1 | Col2 |"
        # Second line is the separator
        assert ":---" in lines[1]

    def test_md_table_left_alignment_marker(self) -> None:
        """Left alignment must produce ':---' in the separator."""
        result = _md_table(["H"], [["v"]], align=["l"])
        separator_line = result.splitlines()[1]
        assert ":---" in separator_line
        assert "---:" not in separator_line

    def test_md_table_right_alignment_marker(self) -> None:
        """Right alignment must produce '---:' in the separator."""
        result = _md_table(["H"], [["v"]], align=["r"])
        separator_line = result.splitlines()[1]
        assert "---:" in separator_line

    def test_md_table_center_alignment_marker(self) -> None:
        """Center alignment must produce ':---:' in the separator."""
        result = _md_table(["H"], [["v"]], align=["c"])
        separator_line = result.splitlines()[1]
        assert ":---:" in separator_line

    def test_md_table_mixed_alignment(self) -> None:
        """Mixed alignment markers are applied column-by-column."""
        result = _md_table(["L", "R", "C"], [["a", "b", "c"]], align=["l", "r", "c"])
        sep = result.splitlines()[1]
        # Expect all three marker styles to appear in the separator row.
        assert ":---" in sep
        assert "---:" in sep
        assert ":---:" in sep

    def test_md_table_default_align_is_left(self) -> None:
        """When align is omitted, every column defaults to left alignment."""
        result = _md_table(["A", "B", "C"], [["1", "2", "3"]])
        sep = result.splitlines()[1]
        # Three left-aligned separators: ":--- | :--- | :---"
        assert sep.count(":---") == 3

    def test_md_table_pipe_in_cell_is_escaped(self) -> None:
        """A pipe character inside a cell value must be escaped as '\\|'."""
        result = _md_table(["H"], [["foo|bar"]])
        data_line = result.splitlines()[2]
        assert "foo\\|bar" in data_line

    def test_md_table_newline_in_cell_is_replaced_with_space(self) -> None:
        """Newline characters inside cell values must be replaced with a space."""
        result = _md_table(["H"], [["line1\nline2"]])
        data_line = result.splitlines()[2]
        assert "line1 line2" in data_line

    def test_md_table_row_count_matches_input(self) -> None:
        """The number of data lines must equal the number of input rows."""
        rows = [["a", "b"], ["c", "d"], ["e", "f"]]
        result = _md_table(["X", "Y"], rows)
        lines = result.splitlines()
        # lines[0] = header, lines[1] = separator, lines[2:] = data
        assert len(lines) - 2 == len(rows)

    def test_md_table_empty_rows_produces_header_and_separator_only(self) -> None:
        """Passing an empty rows list must still yield the header and separator lines."""
        result = _md_table(["Only"], [])
        lines = result.splitlines()
        assert len(lines) == 2
        assert lines[0] == "| Only |"
