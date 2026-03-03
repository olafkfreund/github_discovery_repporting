from __future__ import annotations

"""Excel report renderer — generates a multi-sheet .xlsx workbook."""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")
_PCT_FMT = "0.0%"


def _auto_width(ws: Worksheet) -> None:
    """Set column widths based on content, capped at 60 characters."""
    for col_idx in range(1, (ws.max_column or 1) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            cell = row[0]
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value).split("\n")[0]))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Write a styled header row at row 1."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP


class ExcelRenderer:
    """Generate a multi-sheet .xlsx workbook from report data."""

    def generate_excel(self, report_data: dict, output_path: Path) -> Path:
        wb = Workbook()

        # Sheet 1: Cover
        ws_cover = wb.active
        assert ws_cover is not None
        ws_cover.title = "Cover"
        self._write_cover(ws_cover, report_data)

        # Sheet 2: Executive Summary
        ws_summary = wb.create_sheet("Executive Summary")
        self._write_executive_summary(ws_summary, report_data)

        # Sheet 3: Category Scores
        ws_categories = wb.create_sheet("Category Scores")
        self._write_category_scores(ws_categories, report_data)

        # Sheet 4: Recommendations
        ws_recs = wb.create_sheet("Recommendations")
        self._write_recommendations(ws_recs, report_data)

        # Sheet 5: Benchmarks
        ws_bench = wb.create_sheet("Benchmarks")
        self._write_benchmarks(ws_bench, report_data)

        # Sheet 6: All Findings
        ws_findings = wb.create_sheet("All Findings")
        self._write_findings(ws_findings, report_data)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info("Excel report generated: %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Sheet writers
    # ------------------------------------------------------------------

    def _write_cover(self, ws: Worksheet, data: dict) -> None:
        rows = [
            ("Report Title", data.get("report_title", "")),
            ("Customer", data.get("customer_name", "")),
            ("Organisation", data.get("org_name", "")),
            ("Scan ID", data.get("scan_id", "")),
            ("Generated At", data.get("generated_at", "")),
            ("Overall Score", f"{data.get('overall_score', 0):.1f}%"),
            ("DORA Level", data.get("dora_level", "")),
            ("Platform", data.get("platform_display_name", data.get("platform", ""))),
        ]
        _write_header_row(ws, ["Field", "Value"])
        for row_idx, (field, value) in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=str(value))
        _auto_width(ws)

    def _write_executive_summary(self, ws: Worksheet, data: dict) -> None:
        _write_header_row(ws, ["Section", "Content"])
        row = 2
        ws.cell(row=row, column=1, value="Executive Summary").font = Font(bold=True)
        ws.cell(row=row, column=2, value=data.get("executive_summary", ""))
        ws[f"B{row}"].alignment = _WRAP
        row += 1

        ws.cell(row=row, column=1, value="Overall Maturity").font = Font(bold=True)
        ws.cell(row=row, column=2, value=data.get("overall_maturity", ""))
        ws[f"B{row}"].alignment = _WRAP
        row += 1

        risk_highlights = data.get("risk_highlights", [])
        if risk_highlights:
            ws.cell(row=row, column=1, value="Risk Highlights").font = Font(bold=True)
            ws.cell(row=row, column=2, value="\n".join(f"- {r}" for r in risk_highlights))
            ws[f"B{row}"].alignment = _WRAP

        _auto_width(ws)

    def _write_category_scores(self, ws: Worksheet, data: dict) -> None:
        headers = ["Category", "Score", "Strengths", "Weaknesses"]
        _write_header_row(ws, headers)

        category_scores: dict = data.get("category_scores", {})
        narratives = {n["category"]: n for n in data.get("category_narratives", [])}

        row = 2
        for category, score in category_scores.items():
            narrative = narratives.get(category, {})
            strengths = "\n".join(f"- {s}" for s in narrative.get("strengths", []))
            weaknesses = "\n".join(f"- {w}" for w in narrative.get("weaknesses", []))

            ws.cell(row=row, column=1, value=category)
            score_cell = ws.cell(row=row, column=2, value=score / 100.0)
            score_cell.number_format = _PCT_FMT
            ws.cell(row=row, column=3, value=strengths).alignment = _WRAP
            ws.cell(row=row, column=4, value=weaknesses).alignment = _WRAP
            row += 1

        _auto_width(ws)

    def _write_recommendations(self, ws: Worksheet, data: dict) -> None:
        headers = ["Priority", "Title", "Category", "Effort", "Impact", "Description", "Check IDs"]
        _write_header_row(ws, headers)

        row = 2
        for rec in data.get("recommendations", []):
            ws.cell(row=row, column=1, value=rec.get("priority", ""))
            ws.cell(row=row, column=2, value=rec.get("title", ""))
            ws.cell(row=row, column=3, value=rec.get("category", ""))
            ws.cell(row=row, column=4, value=rec.get("effort", ""))
            ws.cell(row=row, column=5, value=rec.get("impact", ""))
            ws.cell(row=row, column=6, value=rec.get("description", "")).alignment = _WRAP
            check_ids = rec.get("related_check_ids", [])
            ws.cell(row=row, column=7, value=", ".join(check_ids) if check_ids else "")
            row += 1

        _auto_width(ws)

    def _write_benchmarks(self, ws: Worksheet, data: dict) -> None:
        headers = ["Framework", "Level", "Summary"]
        _write_header_row(ws, headers)

        row = 2
        for bench in data.get("benchmark_comparisons", []):
            ws.cell(row=row, column=1, value=bench.get("framework", ""))
            ws.cell(row=row, column=2, value=bench.get("level", ""))
            ws.cell(row=row, column=3, value=bench.get("summary", "")).alignment = _WRAP
            row += 1

        _auto_width(ws)

    def _write_findings(self, ws: Worksheet, data: dict) -> None:
        headers = ["Check ID", "Name", "Category", "Severity", "Status", "Detail", "Score"]
        _write_header_row(ws, headers)

        row = 2
        for f in data.get("all_findings", []):
            ws.cell(row=row, column=1, value=f.get("check_id", ""))
            ws.cell(row=row, column=2, value=f.get("check_name", ""))
            ws.cell(row=row, column=3, value=f.get("category", ""))
            ws.cell(row=row, column=4, value=f.get("severity", ""))
            ws.cell(row=row, column=5, value=f.get("status", ""))
            ws.cell(row=row, column=6, value=f.get("detail", "")).alignment = _WRAP
            ws.cell(row=row, column=7, value=f.get("score", 0))
            row += 1

        _auto_width(ws)
